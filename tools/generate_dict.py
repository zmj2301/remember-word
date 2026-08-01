# -*- coding: utf-8 -*-
"""从 Excel 单词表生成 js/dict.js（含单元字段）
用法：python -X utf8 tools/generate_dict.py
"""
import openpyxl, re, json, os, sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "2026秋季八年级上册英语人教版单词表（整理版）.xlsx")
OUT = os.path.join(BASE, "js", "dict.js")


def clean_word(raw):
    """清理 OCR 垃圾，提取干净英文词/短语"""
    if raw is None:
        return ""
    s = str(raw).strip()
    # 去掉开头的数字 + 'Vocabulary A–Z' 前缀
    s = re.sub(r"^\d+\s*", "", s)
    s = re.sub(r"^Vocabulary\s*A[-–]Z\s*\d*\s*", "", s, flags=re.IGNORECASE)
    # 去掉单个字母+空格前缀（如 'A able'）
    s = re.sub(r"^[A-Z]\s+", "", s)
    # 去掉 '(= xxx)' 变体括号
    s = re.sub(r"\s*[（(]\s*=\s*[^)）]*[)）]\s*", "", s)
    # 去掉 'mm (= mmm)' 残留的 '= ...'
    s = re.sub(r"\s*=\s*[^)\s].*$", "", s)
    # 去掉末尾未闭合括号
    s = re.sub(r"\s*[（(（]\s*$", "", s)
    # 去掉混入词性的行尾（如 'least adv. &'、'face-to-face adj.'）
    s = re.sub(r"\s+(adj\.|n\.|v\.|adv\.|prep\.|pron\.|interj\.|conj\.|&)$", "", s)
    return s.strip()


def clean_pos(raw):
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"\s*[（(].*$", "", s)   # 去掉 (swept...)
    s = re.sub(r"\s+&\s*$", "", s)      # 去掉尾部 &
    return s.strip()


def clean_meaning(raw):
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"^[）;；、，,]+", "", s)   # 去掉开头孤立标点
    return s.strip()


def slug(word):
    s = word.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return re.sub(r"_+", "_", s)


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[wb.sheetnames[0]]

    entries = []
    seen = {}          # cleaned word (lower) -> 已存在
    total = 0

    # 行号范围：跳过 552-584 行的 OCR 字母索引区（内容已在主体或为垃圾）
    skip_start, skip_end = 552, 584

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=5, values_only=True), start=2):
        if skip_start <= idx <= skip_end:
            continue
        raw_word = row[1]
        if not raw_word:
            continue
        unit = str(row[0]).strip() if row[0] else ""
        word = clean_word(raw_word)
        if not word or "Vocabulary" in word:
            continue
        key = word.lower()
        if key in seen:
            continue
        seen[key] = True
        total += 1
        entries.append({
            "unit": unit,
            "word": word,
            "phonetic": (str(row[2]).strip() if row[2] else ""),
            "pos": clean_pos(row[3]),
            "meaning": clean_meaning(row[4]),
        })

    # 按单元顺序分组展示（entries 已按插入顺序保持原序）

    # 生成 dict.js
    lines = []
    lines.append("/**")
    lines.append(" * 词库数据（由 tools/generate_dict.py 从 Excel 生成，勿手改）")
    lines.append(" * 来源：2026秋季八年级上册英语人教版单词表（整理版）.xlsx")
    lines.append(" * 每个词包含：单元 / 单词 / 音标 / 词性 / 中文释义")
    lines.append(" */")
    lines.append("const DICTIONARY = [")

    id_used = {}
    for e in entries:
        eid = slug(e["word"])
        if eid in id_used:
            # id 冲突：加单元前缀（去掉空格），仍冲突则加序号
            prefix = slug(e["unit"]).replace("_", "").replace("unit", "u")
            eid = f"{prefix}_{eid}"
        n = 2
        while eid in id_used:
            eid = f"{prefix}_{slug(e['word'])}_{n}"
            n += 1
        id_used[eid] = True
        lines.append("  {")
        lines.append(f'    id: {json.dumps(eid, ensure_ascii=False)},')
        lines.append(f'    unit: {json.dumps(e["unit"], ensure_ascii=False)},')
        lines.append(f'    word: {json.dumps(e["word"], ensure_ascii=False)},')
        lines.append(f'    phonetic: {json.dumps(e["phonetic"], ensure_ascii=False)},')
        lines.append(f'    pos: {json.dumps(e["pos"], ensure_ascii=False)},')
        lines.append(f'    meaning: {json.dumps(e["meaning"], ensure_ascii=False)}')
        lines.append("  },")
    lines.append("];")
    lines.append("")
    lines.append("window.DICTIONARY = DICTIONARY;")

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    # 统计
    by_unit = {}
    for e in entries:
        by_unit.setdefault(e["unit"], 0)
        by_unit[e["unit"]] += 1
    print("生成完成:", OUT)
    print("总词数:", total)
    for u, c in by_unit.items():
        print(f"  {u}: {c} 词")
    print("缺少音标:", sum(1 for e in entries if not e["phonetic"]))
    print("缺少词性:", sum(1 for e in entries if not e["pos"]))


if __name__ == "__main__":
    main()
